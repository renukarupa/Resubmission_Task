import pytest
import openpyxl
from selenium import webdriver
from page_objects import LoginPage
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from datetime import datetime

EXCEL_PATH = 'test_data.xlsx'

def get_test_data():
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append({
            "test_id": row[0],
            "username": row[1],
            "password": row[2],
        })
    return test_data

def write_test_result(test_id, result):
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == test_id:
            row[4].value = datetime.now().date()
            row[5].value = datetime.now().time()
            row[6].value = result
            break
    workbook.save(EXCEL_PATH)


@pytest.mark.parametrize("data", get_test_data())
def test_login(data):
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    login_page = LoginPage(driver)

    login_page.enter_username(data["username"])
    login_page.enter_password(data["password"])
    login_page.click_login()

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "menu_dashboard_index"))
        )
        result = "Passed"
    except:
        result = "Failed"

    write_test_result(data["test_id"], result)
    driver.quit()
